import os
import time
import json
from openpyxl import Workbook
from openpyxl import load_workbook

# DEF
# Input directory
file_dir = r'/mnt/d/Py_prj/xtData/'
# Output directory
save_dir = r'/mnt/d/Py_prj/xtDataOut/'
if not os.path.exists(save_dir):
    os.makedirs(save_dir)

file_list = os.listdir(file_dir) # File list
file_num = len(file_list) # File number
time_start = time.time() # Start time mark

for i in range(file_num): 
    try:
        time_s =time.time() 
        wbr = load_workbook(file_dir+file_list[i], read_only=True) # Open file read-only
    except FileNotFoundError:
        print('File not found!')
    else:
        wsr = wbr['Sheet1']  # Open workbook Sheet 1
        time_open = time.time()
        print("open file %s pasted %.3fs" %(file_list[i], time_open-time_s))
        wb = Workbook() # New a workbook
        ws = wb.active # Active a worksheet
        num_None = 0
        num_Char0 = 0
        num_264 = 0
        for row in wsr.iter_rows(min_row = 1, min_col = 1, max_col =1): # Handle the Nth column
            for cell in row:
                if cell.value == None :
                    num_None += 1
                elif cell.value == '' :
                    num_Char0 += 1
                else:
                    c = json.loads(str(cell.value)) # Trans strings to dict
                    if c['msgType'] == 264 :
                        num_264 += 1
                    else:
                        # Append data i wanted
                        ws.append([c['endpointSerialCode'], c['softVersion'], c['timestamp'], c['solarLevel'], c['batteryLevel'], c['battleTemperature'], c['temperature']])
        wb.save(save_dir+str(file_list[i])) # Save excel file
        print("NoneCell %d, Char0Cell %d, 264Cell %d " %(num_None, num_Char0, num_264))
        print(file_list[i]+' saved!-----------------------')
        wbr.close() # Close workbook

time_end = time.time()
print("All pasted %.3fs" %(time_end-time_start))
